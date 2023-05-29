"""
!/usr/bin/env python
# -*- coding: utf-8 -*-
@Time    : 2023/5/26 17:58
@Author  : 派大星
@Site    : 
@File    : xmind_to_excel.py
@Software: PyCharm
@desc: xmind转excel
"""
from xmindparser import xmind_to_dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def get_xmind_data(xmind_file_path) -> list:
    """读取xmind文件，返回xmind数据"""
    xmind_data = xmind_to_dict(xmind_file_path)
    return xmind_data


# 解析字典
def parse_dict(_dict: dict):
    title = _dict.get("title", None)
    topics = _dict.get("topics", {})
    return title, topics


def xmind_to_list(data_list: list):
    """将xmind数据转成嵌套列表"""
    case_list = []
    d_topic = data_list[0].get("topic")
    sheet_name = d_topic.get("title")
    d_topics = d_topic.get("topics")
    if d_topics:
        for module_dict in d_topics:
            module_title, module_topics = parse_dict(module_dict)
            if module_topics:
                for scene_dict in module_topics:
                    scene_title, scene_topics = parse_dict(scene_dict)
                    if scene_topics:
                        for case_dict in scene_topics:
                            case_title, case_topics = parse_dict(case_dict)
                            if case_topics:
                                for pre_dict in case_topics:
                                    pre_title, pre_topics = parse_dict(pre_dict)
                                    if pre_topics:
                                        for step_dict in pre_topics:
                                            step_title, step_topics = parse_dict(step_dict)
                                            if step_topics:
                                                for exp_dict in step_topics:
                                                    exp_title, _ = parse_dict(exp_dict)
                                                    case_list.append([module_title, scene_title, case_title, pre_title,
                                                                      step_title, exp_title])
                                            else:
                                                case_list.append([module_title, scene_title, case_title, pre_title,
                                                                  step_title])
                                    else:
                                        case_list.append([module_title, scene_title, case_title, pre_title])
                            else:
                                case_list.append([module_title, scene_title, case_title])
                    else:
                        case_list.append([module_title, scene_title])
            else:
                case_list.append([module_title])

    return sheet_name, case_list


def write_excel(sheet_name, case_list):
    """把嵌套列表数据写入Excel"""
    wb = Workbook()
    ws = wb.active

    ws.title = sheet_name
    sheet = wb[sheet_name]
    sheet_title = ["模块", "场景", "用例", "前置条件", "操作步骤", "期望结果"]
    ws.append(sheet_title)
    for index, _title in enumerate(sheet_title):

        sheet.cell(row=1, column=index+1).fill = PatternFill("solid", fgColor="FFBB02")

    for case_row in case_list:
        ws.append(case_row)

    wb.save(f"{sheet_name}.xlsx")


if __name__ == '__main__':
    data_list = get_xmind_data("../testcase.xmind")
    write_excel(*xmind_to_list(data_list))
